#coding=utf-8
import openpyxl
import os
import time

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.styles import Border, Side, Font #设置字体和边框需要的模块
from openpyxl.writer.excel import ExcelWriter


#判断xlsx文件是否存在，不存在则创建
if os.path.exists('test.xlsx'):
    pass
else:
    # 创建excel
    wb = openpyxl.Workbook()
    # 取得sheet
    ws = wb.worksheets[0]
    # 设置宽高,excel的宽高的单位不同，需要换算
    for i in range(65, 91):  # 调整列宽
        ws.column_dimensions[chr(i)].width = 4.374  # 用chr()将对应的code转为字母
    for i in range(1, 100):  # 调整行高
        ws.row_dimensions[i].height = 27.682
    # 保存excel
    wb.save('test.xlsx')
    time.sleep(2)

#获取xlsx：
wb =load_workbook('test.xlsx')

#获取sheet：
ws = wb.worksheets[0]


a=ws.cell(1,2).value
print(a)

#保存excel文件
wb.save('test.xlsx')