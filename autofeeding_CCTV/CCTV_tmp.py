from openpyxl import load_workbook
from determine import send_keys
import APPinfo
import time
#运行环境开关，正式环境：1   测试环境：2
formal = 2
if formal ==1:
    check_path = "//android.widget.LinearLayout[1]/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.view.View"
    name = 'text("️胖娃娃")'
else:
    check_path = "//android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.view.View"
    name = 'text("笑问天")'

desired_caps = APPinfo.getinfo()
driver = APPinfo.getdriver()
time.sleep(8)

last_line = 99  #最后行数指定
result = ''
print('收到查询进度请求，准备生成进度报告')
driver.find_element_by_android_uiautomator(name).click()
time.sleep(1)
send_keys(driver,'生成的report将以逐行扫描的方式发送给您,分别为：\n概要\n详细\n实施状态\n请稍候\n（试用版，后期优化报告生成逻辑及发送方式优化）')
for x in range(2,last_line):
    for y in range(1,4):
        work =load_workbook('autofeeding_CCTV_plan.xlsx')
        work_sheet = work['Sheet1']
        cell_text = work_sheet.cell(x,y).value
        if cell_text != None:
            print(cell_text)
            send_keys(driver,cell_text)
        else:
            send_keys(driver, '进度报告发送完毕')
            break
    if cell_text != None:
        pass
    else:
        break

    # #追加判断当前单元格是否有值，没有值则跳过
    # if date_before == None:
    #     print('当前单元格：'+str(x)+','+str(y)+'未检测到数据，请检查原文件格式')
    #     continue
    # else:
    #     pass
    #
    # pattern = re.compile(r'\d+')
    # date_list = pattern.findall(date_before)
    # Y = date_list[0]
    # M = date_list[1]
    # D = date_list[2]
    #
    # #判断取得的数字长度
    # if len(Y) ==4 :
    #     pass
    # else :
    #     exit('年月日格式异常')
    # if len(M) == 2 :
    #     pass
    # else :
    #     M ='0'+M
    # if len(D) == 2 :
    #     pass
    # else :
    #     D ='0'+D
    #
    # #整合字符串
    # date_after = Y+M+D
    #
    # # 写数据
    # y = 2
    # work_sheet.cell(x, y).value = date_after
    # work.save('work.xlsx')
driver.keyevent('4')  # backkey,返回主画面