#coding=utf-8
import openpyxl
import os
import time
import random
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment,Font #设置字体,居中和边框需要的模块
from openpyxl.writer.excel import ExcelWriter

#定义预填数字区域位置生成的方法
def preset():
    set_list = [0,0,0,0,0,0,0,0,0]        #创建列表,默认值都为0，代表不填数字
    list_tmp = []
    i = 0
    while i < 5:                #生成5个不重复数字，代表代表列表中的位置
        a = random.randint(0,8)
        if a not in list_tmp:
            list_tmp.append(a)
            i += 1
        else:
            pass
    for i in list_tmp:      #根据生成的5个数字代表的位置，将值改为 1 ，代表写入预填数字
        set_list[i] = 1
    return set_list             #返回修改后的随机位置

#定义生成5X9列表的方法，确定空值区域用，子list内不可重复
def list5X9():
    #定义生成5个不重复随机数字的方法
    def random_num_5():
        i = 0
        list_tmp = []
        while i < 5:                #生成5个不重复数字
            a = random.randint(1,9)
            if a not in list_tmp:
                list_tmp.append(a)
                i += 1
            else:
                pass
        return list_tmp
    global list_complete, list_tmp
    # 对所有list初始化
    list_complete = []
    list_tmp = []
    for i in range(9):
        list_tmp = random_num_5()
        list_complete.append(list_tmp)
    return list_complete

#定义生成9X9列表的方法，用于填写数字，结果需符合数独规则
def list9X9():
    con = 0  # 统计用变量
    msg_display = 'OFF'  # 信息显示开关
    # 定义生成9个随机数字的方法
    def random_num_9():
        i = 0
        list_tmp = []
        while i < 9:  # 生成９个不重复数字
            a = random.randint(1, 9)
            if a not in list_tmp:
                list_tmp.append(a)
                i += 1
            else:
                pass
        return list_tmp
    global list_complete,list1,list2,list3,list4,list5,list6,list7,list8,list9
    # 对所有list初始化
    list_complete =[]
    list1 = []
    list2 = []
    list3 = []
    list4 = []
    list5 = []
    list6 = []
    list7 = []
    list8 = []
    list9 = []
    while True:
        if len(list_complete) == 9:
            break
        else:
            # 生成子list1
            list1 = random_num_9()
            if msg_display == 'ON':
                print('生成的list1:', list1)

            # 生成子list2
            while True:
                try:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0
                    c = 0  # 当前位数判断
                    while True:
                        if len(list_tmp) == 0:
                            break
                        elif list_tmp[i] not in list1[0:3] and c <= 2:
                            list_x.append(list_tmp[i])
                            list_tmp.pop(i)
                            c += 1
                            i = 0
                        elif list_tmp[i] not in list1[3:6] and 3 <= c <= 5:
                            list_x.append(list_tmp[i])
                            list_tmp.pop(i)
                            c += 1
                            i = 0
                        elif list_tmp[i] not in list1[6:] and 6 <= c <= 8:
                            list_x.append(list_tmp[i])
                            list_tmp.pop(i)
                            c += 1
                            i = 0
                        else:
                            i += 1
                    break
                except:
                    if len(list_x) == 9:
                        break
                    else:
                        pass
                    pass
            list2 = list_x
            if msg_display == 'ON':
                print('生成的list2:', list2)

            # 生成子list3
            while True:
                list_x = []
                list_tmp = random_num_9()  # 创建一个临时列表
                i = 0
                c = 0  # 当前位数判断
                a = 0  # 统计循环次数
                try:
                    while True:
                        if len(list_tmp) == 0:
                            break
                        elif list_tmp[i] not in (list1[0:3] + list2[0:3]) and c <= 2:
                            list_x.append(list_tmp[i])
                            list_tmp.pop(i)
                            c += 1
                            i = 0
                        elif list_tmp[i] not in (list1[3:6] + list2[3:6]) and 3 <= c <= 5:
                            list_x.append(list_tmp[i])
                            list_tmp.pop(i)
                            c += 1
                            i = 0
                        elif list_tmp[i] not in (list1[6:] + list2[6:]) and 6 <= c <= 8:
                            list_x.append(list_tmp[i])
                            list_tmp.pop(i)
                            c += 1
                            i = 0
                        else:
                            if a <= 81:
                                i += 1
                                a += 1
                                pass
                            else:
                                break
                    break
                except:
                    if len(list_x) == 9:
                        break
                    else:
                        pass
            list3 = list_x
            if msg_display == 'ON':
                print('生成的list3:', list3)

            # 生成子list4
            list_x = []  # 初始化list_x
            while True:
                if len(list_x) == 9:
                    break
                else:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0  # 临时列表的下标
                    c = 0  # 当前位数判断
                    a = 0  # 统计循环次数
                    try:
                        while True:
                            if len(list_tmp) == 0:
                                break
                            elif list_tmp[i] not in (list1[c], list2[c], list3[c]):
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            else:
                                if a <= 999:
                                    i += 1
                                    a += 1
                                else:
                                    break
                        break
                    except:
                        if len(list_x) == 9:
                            break
                        else:
                            pass
            list4 = list_x
            if msg_display == 'ON':
                print('生成的list4:', list4)

            # 生成子list5
            list_x = []  # 初始化list_x
            while True:
                if len(list_x) == 9:
                    break
                else:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0  # 临时列表的下标
                    c = 0  # 当前位数判断
                    a = 0  # 统计循环次数
                    try:
                        while True:
                            # print('*list5*当前临时列表',list_tmp)
                            if len(list_tmp) == 0:
                                break
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list4[0], list4[1], list4[2]) and c <= 2:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list4[3], list4[4], list4[5]) and 3 <= c <= 5:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list4[6], list4[7], list4[8]) and 6 <= c <= 8:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            else:
                                if a <= 999:
                                    i += 1
                                    a += 1
                                else:
                                    break
                        break
                    except:
                        if len(list_x) == 9:
                            break
                        else:
                            pass
            list5 = list_x
            if msg_display == 'ON':
                print('生成的list5:', list5)

            # 生成子list6
            list_x = []  # 初始化list_x
            con = 0  # 初始化统计用变量
            while True:
                if len(list_x) == 9:
                    break
                else:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0  # 临时列表的下标
                    c = 0  # 当前位数判断
                    a = 0  # 统计循环次数
                    try:
                        while True:
                            # print('*list6*当前临时列表：',list_tmp)
                            # print('*list6*当前处理数据：', list_tmp[i])
                            # print('*list6*当前list_x：', list_x)
                            if len(list_tmp) == 0:
                                break
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list4[0], list4[1], list4[2], list5[0],
                            list5[1], list5[2]) and c <= 2:
                                # print('当前数据', list_tmp[i],'追加成功')
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                # print('当前list_x',list_x)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list4[3], list4[4], list4[5], list5[3],
                            list5[4], list5[5]) and 3 <= c <= 5:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list4[6], list4[7], list4[8], list5[6],
                            list5[7], list5[8]) and 6 <= c <= 8:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            else:
                                if a <= 999:
                                    i += 1
                                    a += 1
                                else:
                                    #print('***********a值超过最大循环***********')
                                    break
                        #print('***********赋值循环完成***********')
                        break
                    except:
                        if len(list_x) == 9:
                            break
                        else:
                            pass
                        # print('***********循环内出现未知错误***********')
                        # print('a=', a)
                        # print('i=', i)
                        # print('--------------')
                        # print(list1)
                        # print(list2)
                        # print(list3)
                        # print(list4)
                        # print(list5)
                        # print('--------------')
                        # print(list_tmp)
                        # print(list_x)
                        con += 1
                        #print('当前为第', con, '次失败')
                        if con == 999:
                            #print('重试超过999次')
                            break
                        else:
                            pass
                        time.sleep(0)
            if len(list_x) < 9:
                #print('list_x生成失败')
                continue
            list6 = list_x
            if msg_display == 'ON':
                print('生成的list6:', list6)

            # 生成子list7
            list_x = []  # 初始化list_x
            while True:
                if len(list_x) == 9:
                    break
                else:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0  # 临时列表的下标
                    c = 0  # 当前位数判断
                    a = 0  # 统计循环次数
                    try:
                        while True:
                            if len(list_tmp) == 0:
                                break
                            elif list_tmp[i] not in (list1[c], list2[c], list3[c], list4[c], list5[c], list6[c]):
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            else:
                                if a <= 999:
                                    i += 1
                                    a += 1
                                else:
                                    break
                        break
                    except:
                        if len(list_x) == 9:
                            break
                        else:
                            pass
            list7 = list_x
            if msg_display == 'ON':
                print('生成的list7:', list7)

            # 生成子list8
            list_x = []  # 初始化list_x
            while True:
                if len(list_x) == 9:
                    break
                else:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0  # 临时列表的下标
                    c = 0  # 当前位数判断
                    a = 0  # 统计循环次数
                    try:
                        while True:
                            # print('*list8*当前临时列表',list_tmp)
                            if len(list_tmp) == 0:
                                break
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list6[c], list7[0], list7[1],
                            list7[2]) and c <= 2:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list6[c], list7[3], list7[4],
                            list7[5]) and 3 <= c <= 5:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list6[c], list7[6], list7[7],
                            list7[8]) and 6 <= c <= 8:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            else:
                                if a <= 999:
                                    i += 1
                                    a += 1
                                else:
                                    break
                        break
                    except:
                        if len(list_x) == 9:
                            break
                        else:
                            pass
            list8 = list_x
            if msg_display == 'ON':
                print('生成的list8:', list8)

            # 生成子list9
            list_x = []  # 初始化list_x
            while True:
                if len(list_x) == 9:
                    break
                else:
                    list_x = []
                    list_tmp = random_num_9()  # 创建一个临时列表
                    i = 0  # 临时列表的下标
                    c = 0  # 当前位数判断
                    a = 0  # 统计循环次数
                    try:
                        while True:
                            # print('*list9*当前临时列表',list_tmp)
                            if len(list_tmp) == 0:
                                break
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list6[c], list7[c], list8[c], list7[0],
                            list7[1], list7[2], list8[0], list8[1], list8[2]) and c <= 2:
                                # print('当前数据', list_tmp[i],'追加成功')
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                # print('当前list_x',list_x)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list6[c], list7[c], list8[c], list7[3],
                            list7[4], list7[5], list8[3], list8[4], list8[5]) and 3 <= c <= 5:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            elif list_tmp[i] not in (
                            list1[c], list2[c], list3[c], list4[c], list5[c], list6[c], list7[c], list8[c], list7[6],
                            list7[7], list7[8], list8[6], list8[7], list8[8]) and 6 <= c <= 8:
                                list_x.append(list_tmp[i])
                                list_tmp.pop(i)
                                c += 1
                                i = 0
                            else:
                                if a <= 999:
                                    i += 1
                                    a += 1
                                else:
                                    break
                        break
                    except:
                        if len(list_x) == 9:
                            break
                        else:
                            pass
            list9 = list_x
            if msg_display == 'ON':
                print('生成的list9:', list9)
        # print('生成的list9:',list9)
        # print('生成的list8:',list8)
        # print('生成的list7:',list7)
        # print('生成的list6:',list6)
        # print('生成的list5:',list5)
        # print('生成的list4:',list4)
        # print('生成的list3:',list3)
        # print('生成的list2:',list2)
        # print('生成的list1:',list1)
        list_complete = []
        list_complete.append(list1)
        list_complete.append(list2)
        list_complete.append(list3)
        list_complete.append(list4)
        list_complete.append(list5)
        list_complete.append(list6)
        list_complete.append(list7)
        list_complete.append(list8)
        list_complete.append(list9)
        #print(list_complete)
        #print('计算结束', list_complete)
        return list_complete
        break

#判断xlsx文件是否存在，不存在则创建
if os.path.exists('sudoku.xlsx'):
    pass
else:
    # 创建excel
    wb = openpyxl.Workbook()
    # 取得sheet
    ws = wb.worksheets[0]
    # 设置宽高,excel的宽高的单位不同，需要换算
    for i in range(65, 91):  # 调整列宽
        ws.column_dimensions[chr(i)].width = 4.374  # 用chr()将对应的code转为字母
    for i in range(1, 100):  # 调整行高
        ws.row_dimensions[i].height = 27.682
    # 保存excel
    wb.save('sudoku.xlsx')
    time.sleep(1)

#获取xlsx：
wb =load_workbook('sudoku.xlsx')
#获取sheet：
ws = wb.worksheets[0]

#定义边框样式，上下左右边框
border_basic = Border(left=Side(border_style='thin',color='000000'),     #定义边框样式:正常，用于分割每个单元格
                right=Side(border_style='thin',color='000000'),
                top=Side(border_style='thin',color='000000'),
                bottom=Side(border_style='thin',color='000000'))
border_right = Border(left=Side(style='thin',color='000000'),     #右侧加粗，用于分割每个完整的九宫格
                      right=Side(style='medium',color='FF000000'),
                      top=Side(style='thin',color='000000'),
                      bottom=Side(style='thin',color='000000'))
border_bottom = Border(left=Side(style='thin',color='000000'),     #底边加粗，用于分割每个完整的九宫格
                       right=Side(style='thin',color='000000'),
                       top=Side(style='thin',color='000000'),
                       bottom=Side(style='medium',color='FF000000'))
border_right_and_bottom = Border(left=Side(style='thin',color='000000'),     #右侧+底边加粗，用于分割每个完整的九宫格
                       right=Side(style='medium',color='FF000000'),
                       top=Side(style='thin',color='000000'),
                       bottom=Side(style='medium',color='FF000000'))
#定义文字居中表示
alignment=Alignment(horizontal='center',
                    vertical='center')
#为9x9区域追加边框
for x in range(1,10):       #9x9的底格
    for y in range(1,10):
        ws.cell(x, y).border = border_basic
        ws.cell(x, y).alignment = alignment
for x in range(1,10):       #每个九宫格的加粗分割线
    for y in range(1,10):
        if y%3 ==0 and y < 9:
            ws.cell(x, y).border = border_right
        else:
            pass
        if x%3 ==0 and x < 9:
            ws.cell(x, y).border = border_bottom
        else:
            pass
        if y%3 ==0 and y < 9 and x%3 ==0 and x < 9:
            ws.cell(x, y).border = border_right_and_bottom
        else:
            pass


'''
规则（第一版）：
1。每个九宫格预填数字为5个，空白区域为4个
2。每个独立九宫格内数字不能重复
3。整个数独行列内数字不能重复，列内数字不能重复

逻辑：
1。生成9*9的二阶list
2。list数据规则：
    子list1为不重复的数字1~9（随机顺序）
    子list2为与已生成list1中相同下标数据不重复，当前下标为0~2时与已生成list1[0~2]的数据不重复，当前下标为3~5时与已生成list1[3~5]的数据不重复，当前下标为6~8时与已生成list1[6~8]的数据不重复
    子list3为与已生成list1，2中相同下标数据不重复，当前下标为0~2时与已生成list1~2[0~2]的数据不重复，当前下标为3~5时与已生成list1~2[3~5]的数据不重复，当前下标为6~8时与已生成list1~2[6~8]的数据不重复
    
    子list4为与已生成list1~3中相同下标数据不重复
    子list5为与已生成list1~4中相同下标数据不重复，当前下标为0~2时与已生成list4[0~2]的数据不重复，当前下标为3~5时与已生成list4[3~5]的数据不重复，当前下标为6~8时与已生成list4[6~8]的数据不重复
    子list6为与已生成list1~5中相同下标数据不重复，当前下标为0~2时与已生成list4~5[0~2]的数据不重复，当前下标为3~5时与已生成list4~5[3~5]的数据不重复，当前下标为6~8时与已生成list4~5[6~8]的数据不重复
    
    子list7为与已生成list1~6中相同下标数据不重复
    子list8为与已生成list1~7中相同下标数据不重复，当前下标为0~2时与已生成list7[0~2]的数据不重复，当前下标为3~5时与已生成list7[3~5]的数据不重复，当前下标为6~8时与已生成list7[6~8]的数据不重复
    子list9为与已生成list1~8中相同下标数据不重复，当前下标为0~2时与已生成list7~8[0~2]的数据不重复，当前下标为3~5时与已生成list7~8[3~5]的数据不重复，当前下标为6~8时与已生成list7~8[6~8]的数据不重复
'''

#将9X9list写入生成的区域
list9X9 =list9X9()          #生成随机的9X9list
for x in range(1,10):       #9x9的底格
    for y in range(1,10):
        ws.cell(x, y).value = list9X9[x-1][y-1]

#对取得的list对应的位置写入空值
list5X9 = list5X9()
for i in range(1,10):
    for a in list5X9[i-1]:
        ws.cell(i, a).value = ''


#保存excel文件
wb.save('sudoku.xlsx')