import time
#运行环境开关，正式环境：1   测试环境：2
formal = 1
if formal ==1:
    check_path = "//android.widget.LinearLayout[1]/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.view.View"
    name = 'text("️胖娃娃")'
else:
    check_path = "//android.widget.LinearLayout[2]/android.widget.LinearLayout/android.widget.LinearLayout[2]/android.widget.LinearLayout[1]/android.view.View"
    name = 'text("笑问天")'

#汇报函数
def master(driver,statuscode):
    '''
    汇报模块
    statuscode：0，系统级错误  1，等待反馈超时  2，反馈信息异常 3.宝宝呼叫主人
    '''
    print('状态异常，开始执行汇报程序')
    error_message = ['系统级错误,请确认监控系统运行情况', '你的宝宝没有发送反馈信息，请确认宝宝状态', '你的宝宝发送的反馈信息异常，请确认宝宝状态', '宝宝呼叫主人']
    warning_text = '警告！！！' + error_message[statuscode]
    driver.find_element_by_android_uiautomator('text("笑问天")').click()
    time.sleep(2)
    driver.find_element_by_id('com.tencent.mm:id/ami').send_keys(warning_text)  # 通过ID选择输入框，并发送文本
    driver.find_element_by_android_uiautomator('text("送信")').click()
    driver.keyevent('4')  # backkey,返回主画面
    print('汇报程序执行完毕')

#反馈监听主函数
def pet_for_main(driver):
    '''
    根据饲养目标的反馈信息给出返回值：
    1 进度查询
    2 呼叫主人
    3 调取监控画面（未完成）
    '''
    print('持续监听中')
    while True:
        wait_command = driver.find_element_by_xpath(check_path).get_attribute("text")  # 饲养对象的返回值
        words =['宝宝你好，自动饲养监控程序已启动。 查询…','进度报告发送完毕', '正在为你呼叫主人，请稍候' , '该功能尚未完成' , '说人话']
        if wait_command in words:  #判断最后显示的信息是否为发送的内容
            pass
        elif wait_command == '1': #1 进度查询
            driver.find_element_by_android_uiautomator(name).click()
            time.sleep(1)
            reporting(driver)
        elif wait_command == '2': #2 呼叫主人
            driver.find_element_by_android_uiautomator(name).click()
            time.sleep(1)
            send_keys(driver, '正在为你呼叫主人，请稍候')
            driver.keyevent('4')  # backkey,返回主画面
            master(driver,statuscode=3)
        elif wait_command == '3': #3 调取监控画面（未完成）
            driver.find_element_by_android_uiautomator(name).click()
            time.sleep(1)
            send_keys(driver,'该功能尚未完成')
            driver.keyevent('4')  # backkey,返回主画面
        else:
            driver.find_element_by_android_uiautomator(name).click()
            time.sleep(1)
            send_keys(driver, '说人话')
            driver.keyevent('4')  # backkey,返回主画面
            break

#时间逻辑反馈函数
def pet_for_time(driver,text):
    '''
    根据饲养目标的反馈信息给出返回值：
    反馈为定义字段，返回值为'safe'
    无反馈或反馈为非定义字段，返回值为'warning'
    '''
    command_list_pet = ['吃了', '刚吃', '吃完了', '恩', '嗯', 'en', '没', '下班了'] #第一期仅匹配简单文本，第二期追加复杂文本判断逻辑及文本管理系统
    no_return_value = 0
    while True:
        wait_command = driver.find_element_by_xpath(check_path).get_attribute("text")  # 饲养对象的返回值
        if wait_command == text and no_return_value < 10:  #判断最后显示的信息是否为text（询问内容），如果没有反馈信息，则等待60秒循环10次
            print('饲养对象没有反馈,继续等待。当前等待次数：'+str(no_return_value))
            no_return_value += 1
            time.sleep(10)
        elif wait_command == text and no_return_value == 10: #等待600秒后仍然没有返回值，则返回warning
            return 1  #等待反馈超时,返回1
            break
        else:
            # 根据饲养目标的反馈信息执行下一步操作
            if wait_command in command_list_pet:
                # 返回值为安全字段
                return 'safe'
                break
            else:
                # 返回值为异常字段
                return 2 #反馈信息异常,返回1
                break

def send_keys(driver,text):
    driver.find_element_by_id('com.tencent.mm:id/ami').send_keys(text)  # 通过ID选择输入框，并发送文本
    driver.find_element_by_android_uiautomator('text("送信")').click()


def reporting(driver):
    from openpyxl import load_workbook
    last_line = 99  # 最后行数指定
    print('收到查询进度请求，准备生成进度报告')
    driver.find_element_by_android_uiautomator(name).click()
    time.sleep(1)
    send_keys(driver, '生成的report将以逐行扫描的方式发送给您,分别为：\n概要\n详细\n实施状态\n请稍候\n（试用版，后期优化报告生成逻辑及发送方式优化）')
    for x in range(2, last_line):
        for y in range(1, 4):
            work = load_workbook('autofeeding_CCTV_plan.xlsx')
            work_sheet = work['Sheet1']
            cell_text = work_sheet.cell(x, y).value
            if cell_text != None:
                print(cell_text)
                send_keys(driver, cell_text)
            else:
                send_keys(driver, '进度报告发送完毕')
                break
        if cell_text != None:
            pass
        else:
            break
    driver.keyevent('4')  # backkey,返回主画面
    time.sleep(1)