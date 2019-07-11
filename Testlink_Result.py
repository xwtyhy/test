#coding=utf-8
import openpyxl
import os
import time

from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter


#判断xlsx文件是否存在，不存在则创建
if os.path.exists('lu.xlsx'):
    pass
else:
    excel = openpyxl.Workbook()
    excel.save('lu.xlsx')
    time.sleep(5)

#获取txt文件：
result = open('lu.txt', mode='r')

#获取xlsx：
print('load xlsx')
excel=load_workbook('lu.xlsx')

#获取sheet：
print('load sheet')
table = excel.worksheets[0]

line = result.readline()
x = 1
while line:
    line = result.readline()  # 读取一行文件，包括换行符
    if 'Y19S12TD-' in line and 'テストケース' in line:
        y =1
        table.cell(row=x, column=y).value = line
    elif 'テスト担当者' in line :
        y = 2
        table.cell(row=x, column=y).value = line
    elif '実施結果:' in line:
        y = 3
        table.cell(row=x, column=y).value = line
    elif '実行に関する注記' in line:
        x += 1
    else:
        pass

#关闭txt文件
result.close()

#保存excel文件
excel.save('lu.xlsx')