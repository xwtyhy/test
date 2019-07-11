from openpyxl import load_workbook
import re

last_line = 11  #需要更改格式的行数+1，例如需要改10行，则输入[11]
date_list = []
for x in range(1,last_line):
    y = 1
    work =load_workbook('work.xlsx')
    work_sheet = work['Sheet1']
    date_before = work_sheet.cell(x,y).value

    #追加判断当前单元格是否有值，没有值则跳过
    if date_before == None:
        print('当前单元格：'+str(x)+','+str(y)+'未检测到数据，请检查原文件格式')
        continue
    else:
        pass

    pattern = re.compile(r'\d+')
    date_list = pattern.findall(date_before)
    Y = date_list[0]
    M = date_list[1]
    D = date_list[2]

    #判断取得的数字长度
    if len(Y) ==4 :
        pass
    else :
        exit('年月日格式异常')
    if len(M) == 2 :
        pass
    else :
        M ='0'+M
    if len(D) == 2 :
        pass
    else :
        D ='0'+D

    #整合字符串
    date_after = Y+M+D

    # 写数据
    y = 2
    work_sheet.cell(x, y).value = date_after
    work.save('work.xlsx')
